#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Генератор импульсов v3‑u6 (модифицирован)
— ... (все предыдущие фичи) ...
— ДОБАВЛЕНО: Загрузка данных из CSV и XLSX (формат как при экспорте)
"""

from __future__ import annotations
import re
from dataclasses import dataclass, field
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import numpy as np
import csv
import os # Для работы с путями и расширениями

try:
    from openpyxl import load_workbook # Используем load_workbook для чтения
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font
    from openpyxl import Workbook # Оставляем Workbook для экспорта
    XLSX_SUPPORT = True
except ImportError:
    XLSX_SUPPORT = False
    print("Warning: openpyxl library not found. XLSX import/export will be disabled.")
    print("Install it using: pip install openpyxl")


# ───── Вспомогательные функции времени (ОБНОВЛЕНО) ─────
# ... (код parse_time, fmt, autoscale, unit_formatter без изменений) ...
# Словарь для ПАРСИНГА (регистронезависимый)
# Ключи: 'f', 'p', 'n', 'u', 'm', 'k', 'meg', 'g', ''
# Добавим поддержку 'µ' как 'u'
_PARSE_UNITS = {'f': 1e-15, 'p': 1e-12, 'n': 1e-9, 'u': 1e-6, 'µ': 1e-6, 'm': 1e-3,
                'k': 1e3, 'meg': 1e6, 'g': 1e9, '': 1}
# Регулярное выражение для парсинга: число, опциональные буквы, опционально 's'
# Позволяет вводить '10n', '5 u', '0.1ms', '100ps', '1k', '2meg', '3gs', '10'
# Буквы приводятся к нижнему регистру для поиска в _PARSE_UNITS
_TIME_RE = re.compile(r'^\s*([+-]?[0-9.]+(?:e[+-]?[0-9]+)?)\s*([a-zµ]*?)s?\s*$', re.IGNORECASE)

def parse_time(s: str) -> float:
    """
    Преобразует строку (e.g., '10n', '5u', '0.1m', '100p', '1k', '2meg', '5') в секунды.
    Регистронезависимо. Распознает f, p, n, u, µ, m, k, meg, g. Число без единицы - секунды.
    """
    s_orig = str(s).strip()
    if not s_orig:
        raise ValueError("Пустая строка времени")

    m = _TIME_RE.fullmatch(s_orig)
    if not m:
        # Если regex не сработал, попробуем просто как float (для чисел без единиц)
        try:
            # Проверим, не является ли строка просто числом (включая научную нотацию)
            # Используем более строгую проверку, чтобы не парсить 'abc' как 0
            try:
                v = float(s_orig)
            except ValueError:
                 # Если это не чистое число, то это ошибка формата
                 raise ValueError(f'Неверный формат времени: "{s_orig}"')

            # Время не может быть отрицательным, но для промежуточных расчетов или напряжения - может
            # Оставим проверку на отрицательность там, где это семантически нужно (TD, TR и т.д.)
            return v
        except ValueError:
            raise ValueError(f'Неверный формат времени: "{s_orig}"')

    v_str, u_str = m.groups()
    unit_lower = u_str.lower()

    try:
        v = float(v_str)
    except ValueError:
        raise ValueError(f'Неверное числовое значение: "{v_str}" в "{s_orig}"')

    # Подбираем множитель из словаря
    multiplier = None
    if unit_lower in _PARSE_UNITS:
        multiplier = _PARSE_UNITS[unit_lower]
    elif unit_lower == 'mega': # Альтернативное написание
        multiplier = _PARSE_UNITS['meg']
    elif unit_lower == 'giga': # Альтернативное написание
        multiplier = _PARSE_UNITS['g']
    elif not unit_lower: # Пустая единица
        multiplier = _PARSE_UNITS['']
    else:
        raise ValueError(f'Неизвестная единица измерения: "{u_str}" в "{s_orig}"')

    result = v * multiplier
    return result

# Словарь для ФОРМАТИРОВАНИЯ (используем 'u', 'M', 'G')
_FMT_UNITS_INV = {1e-15: 'f', 1e-12: 'p', 1e-9: 'n', 1e-6: 'u', 1e-3: 'm',
                  1: 's', 1e3: 'k', 1e6: 'M', 1e9: 'G'}
# Сортированные ключи для поиска диапазона
_FMT_THRESHOLDS = sorted(_FMT_UNITS_INV.keys())

def fmt(sec: float) -> str:
    """
    Форматирует время (секунды) в строку с однобуквенной единицей измерения (f, p, n, u, m, s, k, M, G).
    """
    if sec == 0:
        return '0s'
    a = abs(sec)
    sign = '-' if sec < 0 else '' # Сохраняем знак

    # Ищем подходящий префикс
    best_unit = 's' # По умолчанию
    best_mult = 1.0

    if a >= 1:
        if a >= 1e9: best_unit, best_mult = 'G', 1e-9
        elif a >= 1e6: best_unit, best_mult = 'M', 1e-6
        elif a >= 1e3: best_unit, best_mult = 'k', 1e-3
        else: best_unit, best_mult = 's', 1.0
    elif a > 0:
        if a < 1e-15: # Очень маленькие числа -> научная нотация
             return f'{sec:.3e}s'
        elif a < 1e-12: best_unit, best_mult = 'f', 1e15
        elif a < 1e-9: best_unit, best_mult = 'p', 1e12
        elif a < 1e-6: best_unit, best_mult = 'n', 1e9
        elif a < 1e-3: best_unit, best_mult = 'u', 1e6
        else: best_unit, best_mult = 'm', 1e3
    else: # Отрицательные, но не 0
        # Повторяем логику для модуля, знак добавим в конце
        a_abs = abs(a)
        if a_abs >= 1e9: best_unit, best_mult = 'G', 1e-9
        elif a_abs >= 1e6: best_unit, best_mult = 'M', 1e-6
        elif a_abs >= 1e3: best_unit, best_mult = 'k', 1e-3
        elif a_abs >= 1: best_unit, best_mult = 's', 1.0
        elif a_abs < 1e-15: return f'{sec:.3e}s'
        elif a_abs < 1e-12: best_unit, best_mult = 'f', 1e15
        elif a_abs < 1e-9: best_unit, best_mult = 'p', 1e12
        elif a_abs < 1e-6: best_unit, best_mult = 'n', 1e9
        elif a_abs < 1e-3: best_unit, best_mult = 'u', 1e6
        else: best_unit, best_mult = 'm', 1e3

    # Форматируем число с 4 значащими цифрами
    val_scaled = a * best_mult
    formatted_val = f"{val_scaled:.4g}"

    # Убираем лишние нули и точку для целых чисел в результате форматирования
    if '.' in formatted_val:
        formatted_val = formatted_val.rstrip('0').rstrip('.')

    return f"{sign}{formatted_val}{best_unit}"

def autoscale(total: float) -> tuple[float, str]:
    """
    Определяет множитель и единицу измерения (f, p, n, u, m, s) для оси X.
    Использует секунды для значений >= 1.
    """
    if total <= 0: return 1, 's' # Масштаб 1, единица 's'

    # Используем абсолютное значение для определения масштаба
    a = abs(total)

    if a >= 1: return 1, 's'      # Секунды
    if a >= 1e-3: return 1e3, 'm'      # Миллисекунды
    if a >= 1e-6: return 1e6, 'u'   # Микросекунды
    if a >= 1e-9: return 1e9, 'n'   # Наносекунды
    if a >= 1e-12: return 1e12, 'p'  # Пикосекунды
    # Порог для фемто? или научная нотация?
    # Если total очень мало, например 1e-16
    if a >= 1e-15: return 1e15, 'f'  # Фемтосекунды
    # Если еще меньше, возможно, стоит использовать пико или нано как минимум?
    # Или оставить секунды с научной нотацией? Autoscale обычно для осей...
    # Matplotlib сам справится с малыми числами. Вернем 'f' как самый малый.
    return 1e15, 'f' # Возвращаем фемто для очень малых

def unit_formatter(unit: str):
    """Возвращает форматер оси X для Matplotlib с однобуквенной единицей."""
    # unit теперь это 'f', 'p', 'n', 'u', 'm', 's'.
    unit_str = unit if unit != 's' else ' s' # Добавляем пробел только перед 's'
    # Форматтер для Matplotlib. '{x:g}' автоматически выбирает формат числа.
    return ticker.FuncFormatter(lambda x, pos: f'{x:g}{unit_str}')


# ───── Data class сигнала ─────
@dataclass
class Signal:
    # Основные параметры PULSE
    V0: float; V1: float
    TD: float; TR: float; TF: float; TH: float; TL: float
    N: int;   Name: str
    # Расчетные поля
    tp: float = field(init=False)
    total_time: float = field(init=False)
    scale: float = field(init=False); unit: str = field(init=False)
    # Хранилище для точек, если сигнал был изменен через таблицу или загружен из PWL/CSV/XLSX
    # Если None, используются параметры PULSE. Иначе - эти точки.
    pwl_points: tuple[list[float], list[float]] | None = None

    def __post_init__(self):
        # Если pwl_points уже установлены (например, при загрузке),
        # не вызываем update() сразу, чтобы не сбросить их.
        # Вызовем update() только если pwl_points не заданы.
        if self.pwl_points is None:
            self.update()
        else:
            # Если pwl_points заданы, обновим только total_time и scale/unit
            self._update_time_scale_from_pwl()


    def update(self):
        """Пересчитывает дополнительные параметры сигнала на основе PULSE. Сбрасывает pwl_points."""
        self.TR = max(1e-15, self.TR) # Избегаем нулевых времен для корректного расчета PW
        self.TF = max(1e-15, self.TF)
        self.TH = max(0, self.TH)
        self.TL = max(0, self.TL)
        self.TD = max(0, self.TD)
        self.N = max(0, self.N)
        self.tp = self.TR + self.TH + self.TF + self.TL
        # Убедимся, что tp не ноль, если N>0, чтобы избежать деления на ноль или зависаний
        if self.N > 0 and self.tp <= 1e-15:
            print(f"Warning: Signal '{self.Name}' has N > 0 but calculated period (tp) is near zero. Setting N=0.")
            # Можно установить N=0 или задать минимальный tp
            # self.tp = 1e-9 # Например, 1 нс по умолчанию, если все времена были 0
            self.N = 0 # Альтернатива - обнулить N

        self.total_time = self.TD if self.N == 0 else self.TD + self.N * self.tp
        self.scale, self.unit = autoscale(self.total_time if self.total_time > 0 else 1e-9) # Масштаб по умолчанию если время 0
        # Сбрасываем PWL представление при обновлении параметров
        self.pwl_points = None

    def _update_time_scale_from_pwl(self):
        """Обновляет total_time, scale, unit на основе pwl_points."""
        if self.pwl_points and self.pwl_points[0]:
            self.total_time = self.pwl_points[0][-1]
        else:
            self.total_time = 0
        self.scale, self.unit = autoscale(self.total_time if self.total_time > 0 else 1e-9)


    def set_pwl_points(self, times: list[float], voltages: list[float]):
        """Устанавливает точки PWL и обновляет total_time, scale, unit."""
        if not times:
            self.pwl_points = None
            self.total_time = 0
        else:
            # Убедимся, что время монотонно возрастает и нет дубликатов (кроме обновления напряжения)
            valid_times = []
            valid_voltages = []
            last_t = -float('inf') # Гарантированно меньше первого времени

            for i in range(len(times)):
                t = times[i]
                v = voltages[i]

                if t > last_t:
                    valid_times.append(t)
                    valid_voltages.append(v)
                    last_t = t
                elif t == last_t:
                     # Если время то же, обновляем напряжение последней точки
                    if valid_voltages:
                        valid_voltages[-1] = v
                    # Не обновляем last_t, чтобы следующая точка с таким же временем тоже обновила напряжение
                else: # t < last_t
                    # Игнорируем точки с меньшим временем (ошибка данных)
                    print(f"Warning: Ignored PWL point for signal '{self.Name}' due to non-monotonic time: ({fmt(t)}, {v:g}) after ({fmt(last_t)})")
                    continue

            # Убедимся, что у нас есть хотя бы одна точка
            if not valid_times:
                 print(f"Warning: No valid PWL points found for signal '{self.Name}' after validation.")
                 self.pwl_points = None
                 self.total_time = 0
            else:
                self.pwl_points = (valid_times, valid_voltages)
                # total_time, scale, unit обновятся в _update_time_scale_from_pwl()

        # Обновляем производные поля
        self._update_time_scale_from_pwl()


    def get_waveform_points(self, xmax: float | None = None, force_pulse: bool = False) -> tuple[list[float], list[float]]:
        """
        Генерирует точки (t, v) для построения графика сигнала.
        Использует self.pwl_points если они заданы И force_pulse=False.
        Иначе генерирует из параметров PULSE.
        Если задан xmax, сигнал продолжается до этого значения.
        """
        # --- Выбор источника точек ---
        use_pwl = self.pwl_points is not None and not force_pulse

        if use_pwl:
            # Используем сохраненные точки PWL
            times, voltages = self.pwl_points
            # Копируем, чтобы не изменить оригинал при добавлении xmax
            times = list(times)
            voltages = list(voltages)
            theoretical_end = times[-1] if times else 0
        else:
            # Генерируем точки из параметров PULSE
            times: list[float] = []
            voltages: list[float] = []
            t_cur = 0.0

            # Начальная точка (всегда добавляем 0, V0)
            times.append(0.0)
            voltages.append(self.V0)

            # Задержка (Delay)
            if self.TD > 1e-15: # Только если задержка значима
                t_cur = self.TD # Время конца задержки
                # Добавляем точку перед началом импульсов
                times.append(t_cur)
                voltages.append(self.V0)
            #else: t_cur = 0.0 - уже установлено

            # Импульсы
            if self.N > 0 and self.tp > 1e-15: # Проверяем, что период не нулевой
                for i in range(self.N):
                    # Rise time
                    t_rise_end = t_cur + self.TR
                    if self.TR > 1e-15: # Добавляем точку только если время > 0
                        times.append(t_rise_end)
                        voltages.append(self.V1)
                    else: # Если TR=0, напряжение меняется мгновенно в t_cur
                        if abs(times[-1] - t_cur) < 1e-15: voltages[-1] = self.V1 # Обновляем предыдущую точку
                        else: times.append(t_cur); voltages.append(self.V1) # Или добавляем новую, если время разное
                    t_cur = t_rise_end

                    # High time
                    t_high_end = t_cur + self.TH
                    if self.TH > 1e-15:
                        times.append(t_high_end)
                        voltages.append(self.V1)
                    #else: напряжение V1 уже установлено
                    t_cur = t_high_end

                    # Fall time
                    t_fall_end = t_cur + self.TF
                    if self.TF > 1e-15:
                        times.append(t_fall_end)
                        voltages.append(self.V0)
                    else: # Если TF=0, напряжение меняется мгновенно в t_cur
                        if abs(times[-1] - t_cur) < 1e-15: voltages[-1] = self.V0 # Обновляем предыдущую точку
                        else: times.append(t_cur); voltages.append(self.V0) # Или добавляем новую
                    t_cur = t_fall_end

                    # Low time
                    t_low_end = t_cur + self.TL
                    if self.TL > 1e-15:
                         times.append(t_low_end)
                         voltages.append(self.V0)
                    #else: напряжение V0 уже установлено
                    t_cur = t_low_end # Время на конец периода i

            # Рассчитаем теоретическое время конца последнего импульса/задержки
            theoretical_end = self.TD if self.N == 0 else self.TD + self.N * self.tp

            # Убедимся, что последняя точка соответствует V0 на theoretical_end
            if not times or abs(times[-1] - theoretical_end) > 1e-12:
                 # Добавляем конечную точку, если ее еще нет
                 times.append(theoretical_end)
                 voltages.append(self.V0)
            elif voltages: # Если время совпадает, убедимся, что напряжение = V0
                voltages[-1] = self.V0

        # --- Обработка xmax и очистка (общая для PWL и PULSE) ---
        final_times = []
        final_voltages = []

        # Добавляем точку (0, V0) если ее нет (важно для PWL, где она может отсутствовать)
        if not times or times[0] > 1e-15:
            # Определяем V0 из PWL, если возможно
            start_voltage = voltages[0] if use_pwl and voltages else self.V0
            final_times.append(0.0)
            final_voltages.append(start_voltage)

        # Копируем существующие точки, удаляя дубликаты времени (обновляя напряжение)
        if times:
            last_t = -float('inf')
            for t, v in zip(times, voltages):
                # Игнорируем точки до 0
                if t < -1e-15: continue

                # Округляем очень близкие к 0 значения до 0
                t_proc = 0.0 if abs(t) < 1e-15 else t

                if abs(t_proc - last_t) > 1e-15: # Новое время
                    final_times.append(t_proc)
                    final_voltages.append(v)
                    last_t = t_proc
                else: # То же время, обновляем напряжение
                    if final_voltages:
                        final_voltages[-1] = v
                    # last_t не меняем

        # Обработка xmax: продление сигнала до xmax, если необходимо
        if xmax is not None and theoretical_end < xmax:
            last_v = final_voltages[-1] if final_voltages else self.V0

            # Если последняя точка УЖЕ на theoretical_end (или очень близко)
            if final_times and abs(final_times[-1] - theoretical_end) < 1e-12:
                # Продлеваем до xmax, если он дальше
                if xmax > final_times[-1] + 1e-12:
                    final_times.append(xmax)
                    final_voltages.append(last_v)
            # Если последняя точка НЕ на theoretical_end ИЛИ xmax дальше theoretical_end
            elif xmax > theoretical_end + 1e-12:
                 # Добавляем точку theoretical_end, если ее еще нет ИЛИ она не последняя
                 if not final_times or abs(final_times[-1] - theoretical_end) > 1e-12:
                     final_times.append(theoretical_end)
                     final_voltages.append(last_v)
                 elif final_voltages: # Если последняя точка на theoretical_end, убедимся в правильном напряжении
                      final_voltages[-1] = last_v
                 # Добавим точку xmax
                 final_times.append(xmax)
                 final_voltages.append(last_v)

        # Если после всех манипуляций список пуст, вернем точку (0, V0)
        if not final_times:
            final_times.append(0.0)
            final_voltages.append(self.V0)
            if xmax is not None and xmax > 1e-15:
                final_times.append(xmax)
                final_voltages.append(self.V0)

        return final_times, final_voltages



# ───── Главное приложение ─────
class PulseApp(ttk.Frame):
    FIELDS = (('V0','Ур0'), ('V1','Ур1'), ('TD','Delay'), ('TR','Rise'),
              ('TF','Fall'), ('TH','High'), ('TL','Low'), ('N','Cnt'), ('Name','Имя'))

    def __init__(self, master: tk.Tk):
        super().__init__(master)
        self.pack(fill='both', expand=True)
        master.title('Impulse generator v3‑u6 (с CSV/XLSX импортом)') # Обновил версию
        master.minsize(850, 650) # Увеличим мин. размер
        self.signals: list[Signal] = []
        self._drag_start_index = None
        self.mode = "high-low"
        self._editing_cell_entry = None # Виджет для редактирования ячейки Treeview
        self._style()
        self._layout()
        self._mpl()
        # Флаг, указывающий, что данные для выбранного сигнала взяты из таблицы
        # self.updated_from_table_index = None # Заменено на проверку signal.pwl_points

    # ... (методы _style, _layout, _mpl, on_mode_change без изменений) ...
    def _style(self):
        st = ttk.Style()
        st.theme_use('clam')
        st.configure('TButton', font=('Segoe UI', 10))
        st.configure('TLabel', font=('Segoe UI', 10))
        st.configure('TEntry', font=('Segoe UI', 10))
        st.configure('TLabelframe.Label', font=('Segoe UI', 10, 'bold'))
        # Стиль для Treeview
        st.configure("Treeview.Heading", font=('Segoe UI', 10, 'bold'))
        st.configure("Treeview", font=('Segoe UI', 9), rowheight=22) # Немного уменьшим шрифт строк

    def _layout(self):
        main_paned = ttk.PanedWindow(self, orient='horizontal')
        main_paned.pack(fill='both', expand=True, padx=5, pady=(5, 0))

        # ─── Левая панель (Список сигналов + Таблица точек) ───
        left_panel = ttk.Frame(main_paned, padding=0)
        main_paned.add(left_panel, weight=1) # Меньший вес для левой панели

        # Верхняя часть левой панели: Список сигналов
        list_frame = ttk.Frame(left_panel)
        list_frame.pack(fill='both', expand=True, pady=(0, 5)) # Занимает доступное место

        self.lb = tk.Listbox(list_frame, exportselection=False, font=('Segoe UI', 10), height=8) # Уменьшим высоту списка
        self.lb.pack(side='top', fill='both', expand=True)
        self.lb.bind('<<ListboxSelect>>', self.on_select)
        self.lb.bind('<ButtonPress-1>', self.on_lb_button_press)
        self.lb.bind('<B1-Motion>', self.on_lb_motion)
        self.lb.bind('<Control-Up>', self.move_item_up)
        self.lb.bind('<Control-Down>', self.move_item_down)

        list_buttons_frame = ttk.Frame(list_frame)
        list_buttons_frame.pack(fill='x', pady=(5, 0))
        btn_configs = (('Load', self.load), ('Add', self.add),
                       ('Del', self.delete), ('Export', self.export))
        for i, (text, cmd) in enumerate(btn_configs):
            btn = ttk.Button(list_buttons_frame, text=text, command=cmd)
            btn.pack(side='left', expand=True, fill='x', padx=(0, 2) if i < len(btn_configs)-1 else 0)

        # Разделитель
        ttk.Separator(left_panel, orient='horizontal').pack(fill='x', pady=5)

        # Нижняя часть левой панели: Таблица точек
        table_frame = ttk.LabelFrame(left_panel, text="Точки выбранного сигнала", padding=5)
        table_frame.pack(fill='both', expand=True, pady=(0, 5)) # Занимает оставшееся место

        self.points_table = ttk.Treeview(table_frame, columns=('time', 'voltage'), show='headings', height=6)
        self.points_table.heading('time', text='Время')
        self.points_table.heading('voltage', text='Напряжение')
        self.points_table.column('time', width=100, anchor='e')
        self.points_table.column('voltage', width=80, anchor='e')

        # Скроллбар для таблицы
        points_scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.points_table.yview)
        self.points_table.configure(yscrollcommand=points_scrollbar.set)

        points_scrollbar.pack(side='right', fill='y')
        self.points_table.pack(side='left', fill='both', expand=True)

        # Привязка двойного клика для редактирования
        self.points_table.bind('<Double-1>', self._on_tree_double_click)

        # Кнопка обновления графика по таблице
        self.update_from_table_btn = ttk.Button(left_panel, text='Обновить график по данным таблицы',
                                                command=self._update_graph_from_table, state='disabled')
        self.update_from_table_btn.pack(fill='x', pady=(0, 5))


        # ─── Правая панель (График) ───
        self.plot_fr = ttk.Frame(main_paned, padding=0)
        main_paned.add(self.plot_fr, weight=4) # Больший вес для графика

        # ─── Нижняя панель (Параметры и управление) ───
        editor_frame = ttk.LabelFrame(self, text='Параметры сигнала и графика', padding=10)
        editor_frame.pack(fill='x', padx=5, pady=5)

        # Режим ввода
        mode_frame = ttk.Frame(editor_frame)
        mode_frame.grid(row=0, column=0, columnspan=4, sticky='ew', pady=(0, 10))
        ttk.Label(mode_frame, text="Режим параметров:").pack(side='left')
        self.mode_var = tk.StringVar(value="T_HIGH/T_LOW")
        self.mode_combobox = ttk.Combobox(mode_frame, textvariable=self.mode_var, state="readonly",
                                          values=["T_HIGH/T_LOW", "T_PULSE_WIDTH/T_PERIOD"])
        self.mode_combobox.pack(side='left', padx=5)
        self.mode_combobox.bind("<<ComboboxSelected>>", self.on_mode_change)

        # Поля параметров сигнала
        self.ent: dict[str, ttk.Entry] = {}
        self.param_labels: dict[str, ttk.Label] = {}
        col_param = 0
        for i, (k, lbl) in enumerate(self.FIELDS):
            actual_lbl = lbl
            if k == 'TH': actual_lbl = "High" if self.mode == "high-low" else "PW"
            elif k == 'TL': actual_lbl = "Low" if self.mode == "high-low" else "Period"
            label_widget = ttk.Label(editor_frame, text=actual_lbl+':')
            label_widget.grid(row=i+1, column=col_param, sticky='e', padx=(0,5), pady=2)
            self.param_labels[k] = label_widget
            entry_widget = ttk.Entry(editor_frame, width=12, font=('Segoe UI', 10))
            entry_widget.grid(row=i+1, column=col_param+1, sticky='ew', padx=4, pady=1)
            self.ent[k] = entry_widget

        # Параметры графика
        col_graph = 2
        ttk.Label(editor_frame, text='X‑max:').grid(row=1, column=col_graph, sticky='e', padx=(15,5), pady=2)
        self.xmax_var = tk.StringVar(value='auto')
        ttk.Entry(editor_frame, textvariable=self.xmax_var, width=10, font=('Segoe UI', 10)).grid(row=1, column=col_graph+1, sticky='ew', padx=4, pady=1)

        ttk.Label(editor_frame, text='Шаг тиков X:').grid(row=2, column=col_graph, sticky='e', padx=(15,5), pady=2)
        self.xtick_step_var = tk.StringVar(value='auto')
        self.xtick_step_entry = ttk.Entry(editor_frame, textvariable=self.xtick_step_var, width=10, font=('Segoe UI', 10))
        self.xtick_step_entry.grid(row=2, column=col_graph+1, sticky='ew', padx=4, pady=1)

        # Кнопка Применить / Обновить график (работает с параметрами)
        apply_btn = ttk.Button(editor_frame, text='Применить параметры / Обновить график', command=self.apply_and_draw)
        apply_btn.grid(row=len(self.FIELDS)+1, column=0, columnspan=4, pady=(10, 0), sticky='ew')
        editor_frame.columnconfigure(col_param+1, weight=1)
        editor_frame.columnconfigure(col_graph+1, weight=1)

    def _mpl(self):
        self.fig = plt.Figure(figsize=(6, 4), dpi=100)
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.plot_fr)
        self.canvas_widget = self.canvas.get_tk_widget()
        self.canvas_widget.pack(fill='both', expand=True)
        toolbar_frame = ttk.Frame(self.plot_fr)
        toolbar_frame.pack(fill='x', side='bottom')
        self.toolbar = NavigationToolbar2Tk(self.canvas, toolbar_frame, pack_toolbar=False)
        self.toolbar.update()
        self.toolbar.pack(fill='x')
        self.canvas.draw_idle()

    def on_mode_change(self, event):
        mode_sel = self.mode_var.get()
        new_mode = "pw-period" if mode_sel == "T_PULSE_WIDTH/T_PERIOD" else "high-low"
        if new_mode != self.mode:
            self.mode = new_mode
            # Обновляем метки
            if 'TH' in self.param_labels:
                new_lbl_th = "High" if self.mode == "high-low" else "PW"
                self.param_labels['TH'].config(text=new_lbl_th + ':')
            if 'TL' in self.param_labels:
                new_lbl_tl = "Low" if self.mode == "high-low" else "Period"
                self.param_labels['TL'].config(text=new_lbl_tl + ':')
            # Обновляем значения в полях для текущего выбранного сигнала
            idx = self._get_selected_index()
            if idx is not None and 0 <= idx < len(self.signals):
                # Пересчитываем значения в полях, только если сигнал НЕ на основе PWL
                if self.signals[idx].pwl_points is None:
                     self._update_entries(self.signals[idx])
                # Если на основе PWL, поля все равно показывают базовые параметры,
                # но можно их обновить, если хотим, чтобы они отражали режим
                # self._update_entries(self.signals[idx]) # Раскомментировать, если нужно обновлять поля и для PWL сигналов


    def _read_signal_data(self) -> Signal:
        """Читает данные из полей ввода параметров."""
        v = {k: self.ent[k].get().strip() for k, _ in self.FIELDS}
        try:
            name = v['Name'] if v['Name'] else f'Signal {len(self.signals)+1}'
            # Используем float() напрямую, без 'or 0'/'or 1', чтобы отловить пустые поля как ошибку
            try:
                v0 = float(v['V0'])
                v1 = float(v['V1'])
            except ValueError:
                 raise ValueError("Значения V0 и V1 должны быть числами")

            td = parse_time(v['TD'] or "0")
            tr = parse_time(v['TR'] or "1n") # Используем parse_time для всех времен
            tf = parse_time(v['TF'] or "1n")

            if td < 0: raise ValueError("Время задержки (TD) не может быть отрицательным")
            if tr < 0: raise ValueError("Время нарастания (TR) не может быть отрицательным")
            if tf < 0: raise ValueError("Время спада (TF) не может быть отрицательным")

            if self.mode == "high-low":
                th_str = v['TH'] or "1u"
                tl_str = v['TL'] or "1u"
                th = parse_time(th_str)
                tl = parse_time(tl_str)
                if th < 0: raise ValueError(f"Время High ({th_str}) не может быть отрицательным")
                if tl < 0: raise ValueError(f"Время Low ({tl_str}) не может быть отрицательным")
            else: # Режим T_PULSE_WIDTH/T_PERIOD
                pulse_width_str = v['TH'] or "1u"
                period_str = v['TL'] or "1u"
                pulse_width = parse_time(pulse_width_str)
                period = parse_time(period_str)

                # Проверки для режима PW/Period
                if pulse_width < 0: raise ValueError(f"Ширина импульса PW ({pulse_width_str}) не может быть отрицательной")
                if period < 0: raise ValueError(f"Период ({period_str}) не может быть отрицательным")

                min_pw = tr + tf
                # Добавляем допуск для сравнения с плавающей точкой
                if pulse_width < min_pw - 1e-15:
                    raise ValueError(f"T_PULSE_WIDTH ({fmt(pulse_width)}) должен быть >= TR+TF ({fmt(min_pw)})")
                if period < pulse_width - 1e-15:
                     raise ValueError(f"T_PERIOD ({fmt(period)}) должен быть >= T_PULSE_WIDTH ({fmt(pulse_width)})")

                # Рассчитываем TH и TL для внутреннего хранения
                # Учитываем, что pulse_width может быть очень близок к min_pw
                th = max(0.0, pulse_width - min_pw)
                tl = max(0.0, period - pulse_width)

            # N может быть 0
            n_str = v['N'] or "1"
            try:
                n = int(n_str)
            except ValueError:
                raise ValueError("Число импульсов (Cnt) должно быть целым числом")

            if n < 0: raise ValueError("Число импульсов (Cnt) не может быть отрицательным")

            # Создаем объект Signal. Метод update() будет вызван в __post_init__
            return Signal(V0=v0, V1=v1, TD=td, TR=tr, TF=tf, TH=th, TL=tl, N=n, Name=name)

        except ValueError as e:
            # Добавляем контекст ошибки
            raise ValueError(f"Ошибка в параметрах сигнала: {e}")
        except Exception as e:
            import traceback
            print(traceback.format_exc()) # Логируем полный стектрейс для отладки
            raise ValueError(f"Неожиданная ошибка чтения параметров: {e}")


    def _get_selected_index(self):
        s = self.lb.curselection()
        return int(s[0]) if s else None

    # ... (методы move_in_list, on_lb_button_press, on_lb_motion, move_item_up, move_item_down без изменений) ...
    def move_in_list(self, from_index: int, to_index: int):
        if from_index == to_index: return
        sig = self.signals.pop(from_index)
        self.signals.insert(to_index, sig)
        label = self.lb.get(from_index)
        self.lb.delete(from_index)
        self.lb.insert(to_index, label)
        self.lb.select_clear(0, 'end')
        self.lb.selection_set(to_index)
        self.lb.activate(to_index)
        self.draw() # Перерисовываем, так как порядок изменился

    def on_lb_button_press(self, event):
        self._drag_start_index = self.lb.nearest(event.y)

    def on_lb_motion(self, event):
        cur_index = self.lb.nearest(event.y)
        if cur_index != self._drag_start_index and self._drag_start_index is not None:
             # Проверим, что индексы валидны
             size = self.lb.size()
             if 0 <= self._drag_start_index < size and 0 <= cur_index < size:
                 self.move_in_list(self._drag_start_index, cur_index)
                 self._drag_start_index = cur_index # Обновляем стартовый индекс для продолжения перетаскивания

    def move_item_up(self, event):
        i = self._get_selected_index()
        if i is not None and i > 0:
            self.move_in_list(i, i - 1)
        return "break" # Предотвращаем стандартную обработку

    def move_item_down(self, event):
        i = self._get_selected_index()
        if i is not None and i < self.lb.size() - 1:
            self.move_in_list(i, i + 1)
        return "break"

    def add(self):
        try:
            s = self._read_signal_data()
        except ValueError as e:
            messagebox.showerror('Ошибка ввода', str(e), parent=self.master)
            return

        self.signals.append(s)
        self.lb.insert('end', s.Name)
        self.lb.select_clear(0, 'end')
        self.lb.selection_set('end')
        self.lb.activate('end')
        self.lb.see('end')
        # Сразу обновим таблицу и график
        self.on_select() # Это вызовет _populate_points_table
        self.draw()

    def delete(self):
        i = self._get_selected_index()
        if i is None:
            messagebox.showwarning("Нет выбора", "Сначала выберите сигнал для удаления.", parent=self.master)
            return

        if 0 <= i < len(self.signals):
            confirm = messagebox.askyesno("Подтверждение", f"Удалить сигнал '{self.signals[i].Name}'?", parent=self.master)
            if confirm:
                self.lb.delete(i)
                del self.signals[i] # Используем del для удаления по индексу

                if self.lb.size() > 0:
                    new_selection = min(i, self.lb.size() - 1)
                    self.lb.selection_set(new_selection)
                    self.lb.activate(new_selection)
                    self.on_select() # Обновит панель и таблицу для нового выбора
                else:
                    self._clear_entries()
                    self._clear_points_table() # Очистить таблицу точек
                    self.update_from_table_btn.config(state='disabled')
                self.draw() # Перерисовать график
        else:
             messagebox.showerror("Ошибка", "Не удалось удалить сигнал: неверный индекс.", parent=self.master)


    def apply_and_draw(self):
        """Применяет параметры из полей ввода к выбранному сигналу и перерисовывает."""
        i = self._get_selected_index()
        if i is None:
             messagebox.showwarning("Нет выбора", "Сначала выберите сигнал для применения параметров.", parent=self.master)
             return

        if 0 <= i < len(self.signals):
            try:
                original_name = self.signals[i].Name
                updated_signal_params = self._read_signal_data() # Читаем параметры из полей
                if not updated_signal_params.Name:
                    updated_signal_params.Name = original_name

                # Обновляем объект сигнала, используя новые параметры
                # Это автоматически вызовет update() и сбросит pwl_points
                self.signals[i] = updated_signal_params

                # Если имя изменилось, обновляем Listbox
                if self.lb.get(i) != updated_signal_params.Name:
                    self.lb.delete(i)
                    self.lb.insert(i, updated_signal_params.Name)
                    self.lb.select_set(i)

                # Обновляем таблицу точек, т.к. параметры изменились (и pwl_points сбросились)
                self._populate_points_table(self.signals[i])
                # Делаем кнопку обновления по таблице активной, так как таблица теперь отражает параметры
                self.update_from_table_btn.config(state='normal')

            except ValueError as e:
                messagebox.showerror('Ошибка ввода', str(e), parent=self.master)
                return # Не перерисовываем если ошибка
            except Exception as e:
                import traceback
                messagebox.showerror('Ошибка применения', f"Не удалось применить изменения: {e}\n\n{traceback.format_exc()}", parent=self.master)
                return # Не перерисовываем если ошибка
        # Перерисовываем все графики
        self.draw()


    # --- Методы для работы с таблицей точек ---
    # ... (методы _clear_points_table, _populate_points_table, _on_tree_double_click, _on_edit_save, _on_edit_cancel без изменений) ...
    def _clear_points_table(self):
        """Очищает таблицу точек."""
        if self._editing_cell_entry: # Уничтожаем виджет редактирования, если он есть
             self._editing_cell_entry.destroy()
             self._editing_cell_entry = None
        for item in self.points_table.get_children():
            self.points_table.delete(item)

    def _populate_points_table(self, signal: Signal):
        """Заполняет таблицу точек данными из сигнала."""
        self._clear_points_table()
        # Получаем точки (либо из PWL, либо генерируем из PULSE)
        # Передаем force_pulse=False, чтобы использовать PWL, если они есть
        times, voltages = signal.get_waveform_points(force_pulse=False)

        if not times:
            self.update_from_table_btn.config(state='disabled')
            return

        for i, (t, v) in enumerate(zip(times, voltages)):
            # Форматируем время для отображения
            time_str = fmt(t)
            # Форматируем напряжение (используем 'g' для компактности)
            voltage_str = f"{v:g}"
            # Добавляем строку в Treeview. item id будет использоваться для доступа
            self.points_table.insert('', 'end', iid=str(i), values=(time_str, voltage_str))

        self.update_from_table_btn.config(state='normal')


    def _on_tree_double_click(self, event):
        """Обработчик двойного клика по ячейке таблицы для начала редактирования."""
        # Уничтожаем предыдущий виджет редактирования, если он был
        if self._editing_cell_entry:
            self._editing_cell_entry.destroy()
            self._editing_cell_entry = None

        region = self.points_table.identify_region(event.x, event.y)
        if region != "cell":
            return # Кликнули не по ячейке

        item_id = self.points_table.identify_row(event.y)
        column_id = self.points_table.identify_column(event.x)
        # Проверка column_id, может быть пустым, если клик правее последней колонки
        if not column_id: return
        try:
             column_index = int(column_id.replace('#', '')) - 1 # 0 для времени, 1 для напряжения
        except ValueError:
             return # Не смогли определить индекс колонки

        if not item_id or column_index < 0 or column_index > 1:
            return # Не попали или не та колонка

        # Получаем геометрию ячейки
        x, y, width, height = self.points_table.bbox(item_id, column_id)

        # Получаем текущее значение
        current_values = self.points_table.item(item_id, 'values')
        # Проверка, что values не пустые (может случиться при странных обстоятельствах)
        if not current_values or len(current_values) <= column_index: return
        value = current_values[column_index]

        # Создаем Entry поверх ячейки
        entry = ttk.Entry(self.points_table, font=('Segoe UI', 9))
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, value)
        entry.select_range(0, 'end')
        entry.focus_set()
        self._editing_cell_entry = entry # Сохраняем ссылку на виджет

        # Сохраняем информацию о редактируемой ячейке
        entry.editing_info = {'item_id': item_id, 'column_index': column_index}

        # Привязываем события для завершения редактирования
        entry.bind("<Return>", self._on_edit_save)
        entry.bind("<KP_Enter>", self._on_edit_save) # Enter на цифровой клавиатуре
        entry.bind("<FocusOut>", self._on_edit_save)
        entry.bind("<Escape>", self._on_edit_cancel)

    def _on_edit_save(self, event):
        """Сохраняет измененное значение из Entry в Treeview."""
        entry = event.widget
        if not hasattr(entry, 'editing_info'): return # Не наш виджет

        new_value_str = entry.get()
        info = entry.editing_info
        item_id = info['item_id']
        col_index = info['column_index']

        # Получаем текущие значения строки
        try:
            current_values = list(self.points_table.item(item_id, 'values'))
        except tk.TclError:
            # Элемент мог быть удален, пока редактирование было активно
            entry.destroy()
            self._editing_cell_entry = None
            return

        # Пытаемся валидировать и обновить значение
        try:
            if col_index == 0: # Время
                # Пробуем парсить время, допускаем и просто числа (секунды)
                parsed_time = parse_time(new_value_str)
                if parsed_time < 0: raise ValueError("Время не может быть отрицательным")
                formatted_time = fmt(parsed_time) # Переформатируем для единообразия
                current_values[col_index] = formatted_time
            else: # Напряжение
                parsed_voltage = float(new_value_str)
                current_values[col_index] = f"{parsed_voltage:g}" # Сохраняем как строку

            # Обновляем значения в Treeview
            self.points_table.item(item_id, values=tuple(current_values))

        except ValueError as e:
            messagebox.showerror("Ошибка ввода", f"Неверное значение '{new_value_str}':\n{e}", parent=self.master)
            # Не закрываем редактор при ошибке, даем исправить
            entry.focus_set() # Возвращаем фокус
            return # Прерываем сохранение
        except Exception as e_update:
             print(f"Error updating treeview item: {e_update}")
             # Ошибка обновления элемента (возможно, он был удален)
             # Просто закроем редактор

        # Уничтожаем Entry
        entry.destroy()
        self._editing_cell_entry = None
        # Не обновляем график автоматически, оставляем это на кнопку
        # Но нужно сигнализировать, что данные в таблице изменились и могут отличаться
        # от отображаемого графика (если он был по параметрам)

    def _on_edit_cancel(self, event):
        """Отменяет редактирование и уничтожает Entry."""
        if event.widget == self._editing_cell_entry:
            event.widget.destroy()
            self._editing_cell_entry = None

    def _update_graph_from_table(self):
        """Читает данные из таблицы, обновляет PWL точки сигнала и перерисовывает график."""
        idx = self._get_selected_index()
        if idx is None or not (0 <= idx < len(self.signals)):
             messagebox.showwarning("Нет выбора", "Выберите сигнал, для которого нужно обновить график по таблице.", parent=self.master)
             return

        signal = self.signals[idx]
        new_times = []
        new_voltages = []
        row_num = 0
        last_time = -float('inf') # Для проверки монотонности

        # Проходим по всем строкам таблицы
        item_ids = self.points_table.get_children()
        if not item_ids:
             messagebox.showinfo("Таблица пуста", "Нет точек для обновления.", parent=self.master)
             return

        for item_id in item_ids:
            row_num += 1
            values = self.points_table.item(item_id, 'values')
            # Проверка на случай пустых значений (маловероятно, но возможно)
            if not values or len(values) < 2:
                messagebox.showerror("Ошибка данных в таблице", f"Ошибка в строке {row_num}: Неполные данные.", parent=self.master)
                self.points_table.selection_set(item_id); self.points_table.focus(item_id); self.points_table.see(item_id)
                return

            time_str = values[0]
            voltage_str = values[1]

            try:
                # Парсим время
                t = parse_time(time_str)
                # Парсим напряжение
                v = float(voltage_str)

                # Проверка монотонности времени (важно!)
                if t < last_time - 1e-15: # Допуск для сравнения
                    raise ValueError(f"Время должно монотонно возрастать. Нарушение в строке {row_num} ({fmt(t)} < {fmt(last_time)}).")
                # Если время то же самое, обновляем напряжение последней добавленной точки
                elif abs(t - last_time) < 1e-15 and new_voltages:
                    new_voltages[-1] = v
                    continue # Не добавляем новую точку, только обновляем напряжение
                # Иначе (время больше) - добавляем новую точку
                else:
                     new_times.append(t)
                     new_voltages.append(v)
                     last_time = t

            except ValueError as e:
                messagebox.showerror("Ошибка данных в таблице", f"Ошибка в строке {row_num} (значения: '{time_str}', '{voltage_str}'):\n{e}", parent=self.master)
                # Выделяем ошибочную строку в таблице
                self.points_table.selection_set(item_id)
                self.points_table.focus(item_id)
                self.points_table.see(item_id)
                return # Прерываем обновление

        # Проверка на минимальное количество точек
        if len(new_times) < 1: # Достаточно одной точки (t=0, V)
            messagebox.showerror("Ошибка данных", "Необходимо как минимум 1 точка для определения сигнала.", parent=self.master)
            return
        # Если первая точка не t=0, добавим ее с напряжением первой точки
        if new_times[0] > 1e-15:
            new_times.insert(0, 0.0)
            new_voltages.insert(0, new_voltages[0])


        # Если все успешно прочитано, обновляем PWL данные сигнала
        signal.set_pwl_points(new_times, new_voltages)

        # Обновляем поля V0/V1 в интерфейсе, чтобы отразить начало сигнала из PWL
        # Остальные поля (TR/TF и т.д.) не трогаем, они больше не релевантны для формы сигнала
        self.ent['V0'].delete(0, 'end'); self.ent['V0'].insert(0, f"{new_voltages[0]:g}")
        # Найдем первый отличающийся уровень для V1, или используем V0
        v1_display = new_voltages[0]
        for v in new_voltages:
            if abs(v - v1_display) > 1e-9:
                v1_display = v
                break
        self.ent['V1'].delete(0, 'end'); self.ent['V1'].insert(0, f"{v1_display:g}")


        # Перерисовываем график
        self.draw()
        messagebox.showinfo("Обновление", f"График для сигнала '{signal.Name}' обновлен по данным из таблицы.", parent=self.master)


    # --- Остальные методы ---

    def on_select(self, event=None):
        """Вызывается при выборе сигнала в списке."""
        # Завершаем редактирование ячейки, если оно было активно
        if self._editing_cell_entry:
            # Пытаемся сохранить значение перед сменой сигнала
            self._on_edit_save(tk.Event()) # Имитируем событие для сохранения
            if self._editing_cell_entry: # Если сохранение не удалось (ошибка) и виджет еще есть
                self._editing_cell_entry.destroy()
                self._editing_cell_entry = None

        i = self._get_selected_index()
        if i is None:
            self._clear_entries()
            self._clear_points_table()
            self.update_from_table_btn.config(state='disabled')
            # Если список пуст, график тоже надо очистить
            if not self.signals:
                self.draw()
            return

        if 0 <= i < len(self.signals):
            s = self.signals[i]
            self._update_entries(s) # Обновляем поля параметров
            self._populate_points_table(s) # Заполняем таблицу точек
            # Кнопка "Обновить по таблице" всегда активна, если есть точки в таблице
            self.update_from_table_btn.config(state='normal' if self.points_table.get_children() else 'disabled')
            # Перерисуем график, чтобы снять выделение предыдущего и выделить новый
            self.draw()


    def _update_entries(self, signal: Signal):
         """Обновляет поля ввода параметров на основе данных сигнала."""
         # Если сигнал использует PWL точки, поля V0/V1 берем из них,
         # остальные параметры - из сохраненных базовых параметров сигнала.
         if signal.pwl_points and signal.pwl_points[1]:
             v0_display = signal.pwl_points[1][0]
             v1_display = v0_display
             for v in signal.pwl_points[1]:
                 if abs(v - v0_display) > 1e-9:
                     v1_display = v
                     break
             self.ent['V0'].delete(0, 'end'); self.ent['V0'].insert(0, f"{v0_display:g}")
             self.ent['V1'].delete(0, 'end'); self.ent['V1'].insert(0, f"{v1_display:g}")
         else:
             # Используем V0/V1 из параметров сигнала
             self.ent['V0'].delete(0, 'end'); self.ent['V0'].insert(0, f"{signal.V0:g}")
             self.ent['V1'].delete(0, 'end'); self.ent['V1'].insert(0, f"{signal.V1:g}")

         # Остальные поля всегда показывают сохраненные параметры PULSE
         self.ent['TD'].delete(0, 'end'); self.ent['TD'].insert(0, fmt(signal.TD))
         self.ent['TR'].delete(0, 'end'); self.ent['TR'].insert(0, fmt(signal.TR))
         self.ent['TF'].delete(0, 'end'); self.ent['TF'].insert(0, fmt(signal.TF))
         self.ent['N'].delete(0, 'end'); self.ent['N'].insert(0, signal.N)
         self.ent['Name'].delete(0, 'end'); self.ent['Name'].insert(0, signal.Name)

         # Поля TH/TL зависят от режима
         if self.mode == "high-low":
             self.ent['TH'].delete(0, 'end'); self.ent['TH'].insert(0, fmt(signal.TH))
             self.ent['TL'].delete(0, 'end'); self.ent['TL'].insert(0, fmt(signal.TL))
         else: # Режим PW/Period
             pulse_width = signal.TR + signal.TH + signal.TF
             period = signal.tp # tp = TR + TH + TF + TL
             self.ent['TH'].delete(0, 'end'); self.ent['TH'].insert(0, fmt(pulse_width))
             # Проверим, что период не нулевой перед форматированием
             period_str = fmt(period) if period > 1e-15 else "0s"
             self.ent['TL'].delete(0, 'end'); self.ent['TL'].insert(0, period_str)


    def _clear_entries(self):
        for key in self.ent:
            self.ent[key].delete(0, 'end')

    # ... (метод draw без изменений) ...
    def draw(self):
        """Перерисовывает все графики сигналов."""
        self.fig.clf() # Очищаем фигуру полностью
        if not self.signals:
            # Отображаем сообщение, если нет сигналов
            ax = self.fig.add_subplot(111)
            ax.text(0.5, 0.5, "Нет сигналов для отображения.\nДобавьте или загрузите сигнал.",
                    ha='center', va='center', fontsize=12, color='gray')
            ax.set_xticks([])
            ax.set_yticks([])
            self.canvas.draw_idle()
            self.toolbar.update() # Обновляем панель инструментов
            return

        # --- Получение параметров отображения ---
        xmax_str = self.xmax_var.get().strip().lower()
        xmax_sec_global = None
        try:
            if xmax_str not in ('', 'auto'):
                xmax_sec_global = parse_time(xmax_str)
                if xmax_sec_global <= 0:
                    print("Warning: X-max должен быть > 0. Используется auto.")
                    xmax_sec_global = None
                    self.xmax_var.set('auto')
        except ValueError as e:
            print(f"Warning: Неверное значение X-max '{xmax_str}'. Используется auto. Ошибка: {e}")
            xmax_sec_global = None
            self.xmax_var.set('auto')

        xtick_step_str = self.xtick_step_var.get().strip().lower()
        tick_step_sec = None
        if xtick_step_str not in ('', 'auto'):
             try:
                 tick_step_sec = parse_time(xtick_step_str)
                 if tick_step_sec <= 0:
                     print(f"Warning: Шаг тиков X '{xtick_step_str}' должен быть > 0. Используется auto.")
                     tick_step_sec = None
                     self.xtick_step_var.set('auto')
             except ValueError as e:
                 print(f"Warning: Неверный формат шага тиков X '{xtick_step_str}'. Используется auto. Ошибка: {e}")
                 tick_step_sec = None
                 self.xtick_step_var.set('auto')

        n_signals = len(self.signals)
        idx_sel = self._get_selected_index() # Индекс выбранного в Listbox сигнала

        # Определяем общий масштаб и единицы
        # Если X-max задан глобально, используем его
        # Иначе, находим максимальное время среди всех сигналов
        max_time_for_scale = 0
        if xmax_sec_global is not None:
            max_time_for_scale = xmax_sec_global
        else:
            for s in self.signals:
                 # total_time теперь корректно отражает конец PWL или PULSE
                max_time_for_scale = max(max_time_for_scale, s.total_time)

        # Устанавливаем масштаб по умолчанию, если время нулевое
        if max_time_for_scale <= 0: max_time_for_scale = 1e-9 # e.g., 1 ns

        common_scale, common_unit = autoscale(max_time_for_scale)

        # --- Рисование каждого сигнала ---
        # Используем gridspec для более гибкого управления subplot'ами
        gs = self.fig.add_gridspec(n_signals, 1, hspace=0.1) # Уменьшим расстояние между графиками
        axes = gs.subplots(sharex=True) # Делаем оси X общими
        if n_signals == 1: axes = [axes] # Убедимся, что axes - это всегда список

        for i, (ax, s) in enumerate(zip(axes, self.signals)):
             # ax = self.fig.add_subplot(n_signals, 1, i + 1) # Старый способ

            # Получаем точки сигнала
            # Используем xmax_sec_global, если он задан, иначе None
            # force_pulse=False, чтобы использовать PWL точки если они есть
            times_sec, voltages = s.get_waveform_points(xmax=xmax_sec_global, force_pulse=False)

            # Определяем реальное максимальное время на графике для этого сигнала
            # Если xmax задан, используем его, иначе берем последнюю точку или total_time
            actual_xmax_sec = xmax_sec_global if xmax_sec_global is not None else (times_sec[-1] if times_sec else s.total_time)
            # Убедимся, что xmax не отрицательный и не слишком мал
            if actual_xmax_sec <= 0: actual_xmax_sec = 1e-9

            # Используем общий масштаб и единицы для оси X
            current_scale = common_scale
            current_unit = common_unit

            # Масштабируем время
            times_scaled = [t * current_scale for t in times_sec]

            # Рисуем график
            line_color = 'C0' # Синий по умолчанию
            line_style = '-'
            line_width = 1.5
            # Можно добавить выделение для сигнала, созданного из таблицы/PWL
            # if s.pwl_points is not None:
            #     line_color = 'C1' # Оранжевый
            #     line_style = '--'

            if times_scaled:
                ax.plot(times_scaled, voltages, lw=line_width, color=line_color, ls=line_style)
            else:
                # Если точек нет (например, сигнал с нулевой длительностью)
                # Нарисуем горизонтальную линию на уровне V0
                v0_plot = s.V0 if s.pwl_points is None else s.pwl_points[1][0] # Учитываем PWL
                ax.plot([0, actual_xmax_sec * current_scale], [v0_plot, v0_plot], lw=line_width, color=line_color, ls=line_style)


            # Настройка внешнего вида subplot'а
            title_style = {'fontweight': 'bold', 'color': 'darkblue'} if i == idx_sel else {}
            ax.set_title(s.Name, fontsize=9, loc='left', y=0.85, **title_style) # Сдвинем заголовок чуть ниже
            ax.set_ylabel('V', fontsize=9)
            ax.tick_params(axis='y', labelsize=8) # Уменьшим шрифт тиков Y

            # Настройка пределов по Y
            v_min_data = min(voltages) if voltages else -1.0 # Дефолт, если нет точек
            v_max_data = max(voltages) if voltages else 1.0

            # Используем V0/V1 из PWL, если они есть, для определения диапазона
            v0_eff = s.pwl_points[1][0] if s.pwl_points and s.pwl_points[1] else s.V0
            v1_eff = v0_eff
            if s.pwl_points and s.pwl_points[1]:
                 for v in s.pwl_points[1]:
                     if abs(v - v0_eff) > 1e-9: v1_eff = v; break
            else:
                 v1_eff = s.V1

            v_min_eff = min(v0_eff, v1_eff, v_min_data)
            v_max_eff = max(v0_eff, v1_eff, v_max_data)

            v_range = v_max_eff - v_min_eff
            if abs(v_range) < 1e-9: v_range = max(1.0, abs(v_max_eff)*0.2) # Избегаем нулевого диапазона
            margin = v_range * 0.15
            ax.set_ylim(v_min_eff - margin, v_max_eff + margin)

            # Настройка пределов по X (будет общей из-за sharex)
            xmax_limit_scaled = actual_xmax_sec * current_scale
            # Добавим небольшой отступ справа, если xmax не был задан явно
            xmax_display = xmax_limit_scaled * 1.03 if xmax_sec_global is None else xmax_limit_scaled
            # Убедимся что xmax_display > 0
            if xmax_display <= 0 : xmax_display = 1.0 # Минимальный предел = 1 (в текущем масштабе)

            ax.set_xlim(0, xmax_display)

            # Настройка тиков оси X (делаем только для нижнего графика)
            if i == n_signals - 1:
                if tick_step_sec is not None: # Задан пользователем
                    base_step_scaled = tick_step_sec * current_scale
                    if base_step_scaled > 1e-12: # Проверяем, что шаг не слишком мал для масштаба
                        locator = ticker.MultipleLocator(base=base_step_scaled)
                    else:
                        print(f"Warning: Шаг тиков X ({fmt(tick_step_sec)}) слишком мал для масштаба '{current_unit}'. Используется auto.")
                        locator = ticker.MaxNLocator(nbins='auto', prune='both', integer=False, min_n_ticks=3)
                else: # Автоматический подбор
                    locator = ticker.MaxNLocator(nbins='auto', prune='both', integer=False, min_n_ticks=3)

                ax.xaxis.set_major_locator(locator)
                ax.xaxis.set_major_formatter(unit_formatter(current_unit)) # Используем общий unit
                ax.set_xlabel(f'Время ({current_unit})', fontsize=9)
                ax.tick_params(axis='x', labelsize=8) # Уменьшим шрифт тиков X
            else:
                 # Скрываем тики и метки для верхних графиков (не нужно из-за sharex)
                 # ax.set_xticklabels([]) # Уже делается через sharex
                 pass


            # Сетка
            ax.grid(True, ls=':', lw=0.6, color='lightgrey')


        # --- Финальная настройка и отрисовка ---
        # fig.align_ylabels(axes) # Выравниваем метки Y (если matplotlib >= 3.1)
        try:
            # self.fig.tight_layout(h_pad=0.1) # Tight layout может быть лучше с sharex
            # Используем constrained_layout для лучшего распределения места
             self.fig.set_layout_engine('constrained')
        except Exception as e_layout:
             print(f"Layout engine failed: {e_layout}. Layout may be suboptimal.")
             try: self.fig.tight_layout(h_pad=0.1) # Запасной вариант
             except: pass

        self.canvas.draw_idle() # Запрос на перерисовку в основном цикле Tk
        self.toolbar.update() # Обновляем панель инструментов Matplotlib


    # --- Методы загрузки/экспорта ---
    # ... (метод _choose_export_format без изменений) ...
    def _choose_export_format(self) -> tuple[str | None, bool | None]:
        """
        Создает диалоговое окно для выбора формата экспорта и источника данных.
        Возвращает кортеж: (выбранный_формат, использовать_данные_таблицы)
        или (None, None) при отмене.
        """
        dialog = tk.Toplevel(self.master)
        dialog.title("Параметры экспорта")
        dialog.transient(self.master)
        dialog.grab_set()
        dialog.resizable(False, False)
        # Центрирование относительно главного окна
        x = self.master.winfo_rootx() + (self.master.winfo_width() // 2) - 150 # Примерная ширина окна / 2
        y = self.master.winfo_rooty() + (self.master.winfo_height() // 2) - 100 # Примерная высота окна / 2
        dialog.geometry(f"300x200+{x}+{y}")

        formats = ["PULSE", "PWL", "CSV"]
        if XLSX_SUPPORT:
            formats.append("XLSX")

        # --- Выбор формата ---
        format_frame = ttk.Frame(dialog, padding=5)
        format_frame.pack(fill='x', padx=10, pady=(10, 5))
        ttk.Label(format_frame, text="Формат экспорта:").pack(side='left')
        format_var = tk.StringVar(value=formats[0]) # Выбираем первый по умолчанию
        format_combo = ttk.Combobox(format_frame, textvariable=format_var, values=formats, state='readonly', width=15)
        format_combo.pack(side='right', padx=(5, 0))
        # Устанавливаем фокус на комбобокс
        format_combo.focus_set()


        # --- Выбор источника данных ---
        source_frame = ttk.Frame(dialog, padding=5)
        source_frame.pack(fill='x', padx=10, pady=5)
        self.export_use_table_var = tk.BooleanVar(value=False) # По умолчанию - по параметрам
        source_check = ttk.Checkbutton(source_frame,
                                       text="По точкам из таблицы (иначе по параметрам)",
                                       variable=self.export_use_table_var)
        source_check.pack(anchor='w')

        # --- Кнопки OK/Отмена ---
        btn_frame = ttk.Frame(dialog, padding=(0, 10))
        btn_frame.pack(fill='x', side='bottom') # Помещаем кнопки вниз

        selected_format = None
        export_from_table = None

        def on_ok():
            nonlocal selected_format, export_from_table
            selected_format = format_var.get()
            export_from_table = self.export_use_table_var.get()
            dialog.destroy()

        def on_cancel():
            nonlocal selected_format, export_from_table
            selected_format = None
            export_from_table = None
            dialog.destroy()

        # Размещаем кнопки по центру
        center_frame = ttk.Frame(btn_frame)
        center_frame.pack()

        ok_btn = ttk.Button(center_frame, text="OK", command=on_ok, width=10)
        ok_btn.pack(side='left', padx=5)
        cancel_btn = ttk.Button(center_frame, text="Отмена", command=on_cancel, width=10)
        cancel_btn.pack(side='left', padx=5)

        # Биндинги для Enter/Escape
        dialog.bind('<Return>', lambda e: ok_btn.invoke()) # Имитируем нажатие OK
        dialog.bind('<Escape>', lambda e: cancel_btn.invoke()) # Имитируем нажатие Отмена

        dialog.wait_window()
        return selected_format, export_from_table

    # ... (метод export без изменений) ...
    def export(self):
        """Экспортирует данные сигналов в выбранный файл."""
        if not self.signals:
            messagebox.showinfo('Экспорт', 'Нет сигналов для экспорта.', parent=self.master)
            return

        format_choice, use_table_data = self._choose_export_format()
        if format_choice is None: # Пользователь отменил
            return

        # --- Особая логика для PULSE формата ---
        # PULSE формат всегда основан на параметрах, игнорируем выбор источника
        force_pulse_export = False
        data_source_msg = "" # Сообщение для пользователя об источнике данных

        if format_choice == "PULSE":
            if use_table_data:
                 # Предупреждаем, что выбор игнорируется
                 messagebox.showwarning("Экспорт PULSE", "Экспорт в PULSE всегда использует параметры сигнала.\nВыбор 'По точкам из таблицы' будет проигнорирован.", parent=self.master)
            force_pulse_export = True # Принудительно используем параметры для PULSE
            data_source_msg = "Параметры (PULSE формат)"
        else:
            force_pulse_export = not use_table_data # Если НЕ из таблицы, то принудительно по параметрам
            data_source_msg = "Таблица" if use_table_data else "Параметры"

        # Получаем X-max для экспорта (если задан)
        xmax_str = self.xmax_var.get().strip().lower()
        xmax_sec = None
        if xmax_str not in ('', 'auto'):
            try:
                xmax_sec = parse_time(xmax_str)
                if xmax_sec <= 0: xmax_sec = None
            except ValueError: xmax_sec = None

        # Определяем расширение файла и типы файлов
        file_types = [("All files", "*.*")]
        default_name_part = 'params' if force_pulse_export else 'table'
        if format_choice == "CSV":
            def_ext = ".csv"
            types = [("CSV files", "*.csv")]
        elif format_choice == "XLSX":
            def_ext = ".xlsx"
            types = [("Excel files", "*.xlsx")]
        else: # PULSE, PWL
            def_ext = ".txt"
            types = [("Text files", "*.txt"), ("PULSE files", "*.pulse"), ("PWL files", "*.pwl")]

        file_types.insert(0, types[0]) # Помещаем специфичный тип первым

        filepath = filedialog.asksaveasfilename(
            defaultextension=def_ext,
            filetypes=file_types,
            initialfile=f"signals_{format_choice.lower()}_{default_name_part}{def_ext}", # Предлагаем имя файла
            title=f"Сохранить как {format_choice} файл (Источник: {data_source_msg})",
            parent=self.master
        )
        if not filepath: return

        try:
            if format_choice == "PWL":
                with open(filepath, 'w', encoding='utf-8') as f:
                    for s in self.signals:
                        # Получаем точки с учетом выбора источника и xmax
                        times, voltages = s.get_waveform_points(xmax_sec, force_pulse=force_pulse_export)
                        # Используем научную нотацию для времени для большей точности
                        # Используем 'g' для напряжения
                        points_str = " ".join(f"{t:.17g} {v:g}" for t, v in zip(times, voltages))
                        f.write(f"{s.Name}:PWL({points_str})\n")
                messagebox.showinfo('Экспорт завершен', f'Сигналы сохранены в PWL файл:\n{filepath}\n(Источник: {data_source_msg})', parent=self.master)

            elif format_choice == "PULSE":
                # Здесь всегда используем параметры (force_pulse_export = True)
                with open(filepath, 'w', encoding='utf-8') as f:
                    for s in self.signals:
                         # Используем внутренние параметры TR, TH, TF, TL, TD, N, V0, V1
                         # ВАЖНО: Формат PULSE ожидает TH и TP (период), а не TH и TL
                         tp_val = s.TR + s.TH + s.TF + s.TL
                         # Используем 'g' для напряжений и N, научную нотацию для времен
                         f.write(f"{s.Name}:PULSE({s.V0:g} {s.V1:g} {s.TD:.17g} {s.TR:.17g} {s.TF:.17g} {s.TH:.17g} {tp_val:.17g} {s.N})\n")
                messagebox.showinfo('Экспорт завершен', f'Сигналы сохранены в PULSE файл:\n{filepath}\n(Источник: {data_source_msg})', parent=self.master)

            elif format_choice in ["CSV", "XLSX"]:
                # 1. Собрать все уникальные временные точки всех сигналов
                all_times = set()
                signals_points_map = {}
                for s in self.signals:
                    # Получаем точки с учетом выбора источника и xmax
                    times, voltages = s.get_waveform_points(xmax_sec, force_pulse=force_pulse_export)
                    # Используем небольшой допуск при добавлении времен в set
                    # чтобы избежать дубликатов из-за ошибок округления
                    for t in times:
                        # Можно округлить до разумного числа знаков (e.g., 15-17)
                        # all_times.add(round(t, 17))
                        # Или просто использовать как есть, set разберется
                         all_times.add(t)

                    signals_points_map[s.Name] = (times, voltages)

                sorted_times = sorted(list(all_times))
                # Убедимся, что 0.0 есть в списке, если его там нет
                if 0.0 not in all_times and (not sorted_times or sorted_times[0] > 1e-18):
                    sorted_times.insert(0, 0.0)


                # 2. Создать строки данных
                headers = ['Time (s)'] + [s.Name for s in self.signals]
                data_rows = [headers]
                # Создаем lookup для быстрого поиска индекса времени
                time_index_map = {t: i for i, t in enumerate(sorted_times)}

                # Готовим интерполированные данные для каждого сигнала
                signal_interp_voltages = {}
                for s in self.signals:
                    s_times, s_voltages = signals_points_map[s.Name]
                    # Используем numpy для быстрой интерполяции (step-функция)
                    # 'previous' означает, что значение сохраняется до следующей точки
                    interp_func = np.interp
                    # Создаем numpy массивы
                    np_s_times = np.array(s_times)
                    np_s_voltages = np.array(s_voltages)
                    np_sorted_times = np.array(sorted_times)

                    # Интерполируем значения для всех времен в sorted_times
                    # np.interp делает линейную интерполяцию. Нам нужна ступенчатая.
                    # Найдем индексы в s_times для каждого времени в sorted_times
                    # searchsorted возвращает индекс i, такой что s_times[i-1] <= t < s_times[i]
                    indices = np.searchsorted(np_s_times, np_sorted_times, side='right')
                    # Берем значение из предыдущей точки s_times
                    indices = np.maximum(0, indices - 1) # Убедимся, что индекс не отрицательный
                    interp_voltages = np_s_voltages[indices]

                    # Особый случай для времени t=0, если оно не было в исходных данных сигнала
                    if 0.0 not in np_s_times and 0.0 in time_index_map:
                         start_v = s.V0 if s.pwl_points is None else s.pwl_points[1][0]
                         interp_voltages[time_index_map[0.0]] = start_v


                    signal_interp_voltages[s.Name] = interp_voltages

                # Собираем строки
                for i, t in enumerate(sorted_times):
                    row = [f"{t:.17g}"] # Время с высокой точностью
                    for s in self.signals:
                        v = signal_interp_voltages[s.Name][i]
                        row.append(f"{v:g}") # Напряжение компактно
                    data_rows.append(row)


                # 3. Записать в файл
                if format_choice == "CSV":
                    with open(filepath, 'w', newline='', encoding='utf-8') as f:
                        writer = csv.writer(f, delimiter=',') # Стандартный разделитель для CSV
                        writer.writerows(data_rows)
                    messagebox.showinfo('Экспорт завершен', f'Временные диаграммы сохранены в CSV файл:\n{filepath}\n(Источник: {data_source_msg})', parent=self.master)
                else: # XLSX
                    if not XLSX_SUPPORT: # Доп. проверка, хотя диалог не должен был это позволить
                         raise ImportError("Библиотека openpyxl не найдена, не могу сохранить XLSX.")
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Signals"
                    for r_idx, row_data in enumerate(data_rows, 1):
                        # Преобразуем строки с числами в числа для Excel
                        converted_row = []
                        for cell_idx, cell_value in enumerate(row_data):
                            if r_idx > 1: # Не трогаем заголовки
                                try:
                                    # Первую колонку (время) и остальные (напряжение) пробуем как float
                                    converted_row.append(float(cell_value))
                                except ValueError:
                                    converted_row.append(cell_value) # Оставляем как есть, если не число
                            else:
                                converted_row.append(cell_value)
                        ws.append(converted_row)

                        # Делаем заголовки жирными
                        if r_idx == 1:
                            for col in range(1, len(headers) + 1):
                                ws.cell(row=1, column=col).font = Font(bold=True)
                    # Автоподбор ширины колонок
                    for col in ws.columns:
                        max_length = 0
                        column_letter = get_column_letter(col[0].column)
                        for cell in col:
                            try:
                                if cell.value is not None:
                                    # Форматируем числа для корректного измерения длины
                                    if isinstance(cell.value, (int, float)):
                                        cell_str = f"{cell.value:g}" # Используем 'g' для компактности
                                    else:
                                        cell_str = str(cell.value)

                                    if len(cell_str) > max_length:
                                        max_length = len(cell_str)
                            except: pass
                        # Устанавливаем ширину с небольшим запасом
                        adjusted_width = (max_length + 2) * 1.1
                        ws.column_dimensions[column_letter].width = adjusted_width

                    wb.save(filepath)
                    messagebox.showinfo('Экспорт завершен', f'Временные диаграммы сохранены в Excel файл:\n{filepath}\n(Источник: {data_source_msg})', parent=self.master)

        except ImportError as e:
             messagebox.showerror("Ошибка экспорта", f"Не удалось экспортировать в XLSX:\n{e}\nУстановите библиотеку 'openpyxl'.", parent=self.master)
        except Exception as e:
            import traceback
            messagebox.showerror("Ошибка экспорта", f"Не удалось сохранить файл:\n{e}\n\n{traceback.format_exc()}", parent=self.master)


    def load(self):
        """Загружает сигналы из выбранного файла (PULSE/PWL txt, CSV, XLSX)."""
        file_types = [
            ("Поддерживаемые файлы", "*.txt *.pulse *.pwl *.csv *.xlsx"),
            ("Text files (PULSE/PWL)", "*.txt *.pulse *.pwl"),
            ("CSV files", "*.csv"),
        ]
        if XLSX_SUPPORT:
            file_types.append(("Excel files", "*.xlsx"))
        file_types.append(("All files", "*.*"))

        filepath = filedialog.askopenfilename(
            filetypes=file_types,
            title="Загрузить сигналы из файла",
            parent=self.master
        )
        if not filepath: return

        loaded_signals = []
        parse_errors = []
        file_type = None
        file_ext = os.path.splitext(filepath)[1].lower()

        # Определяем тип файла
        if file_ext in ['.txt', '.pulse', '.pwl']:
            file_type = 'text'
        elif file_ext == '.csv':
            file_type = 'csv'
        elif file_ext == '.xlsx':
            if not XLSX_SUPPORT:
                messagebox.showerror("Ошибка загрузки", "Библиотека 'openpyxl' не найдена. Невозможно загрузить XLSX файлы.\nУстановите ее: pip install openpyxl", parent=self.master)
                return
            file_type = 'xlsx'
        else:
            # Попробуем определить как текстовый, если расширение неизвестно
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    f.read(10) # Пробуем прочитать немного
                file_type = 'text' # Похоже на текст
                print(f"Warning: Неизвестное расширение '{file_ext}', пытаемся обработать как текстовый файл.")
            except Exception:
                 messagebox.showerror("Ошибка загрузки", f"Неподдерживаемый тип файла или ошибка чтения: {os.path.basename(filepath)}", parent=self.master)
                 return

        # --- Загрузка в зависимости от типа ---
        try:
            if file_type == 'text':
                loaded_signals, parse_errors = self._load_text(filepath)
            elif file_type == 'csv':
                loaded_signals, parse_errors = self._load_csv(filepath)
            elif file_type == 'xlsx':
                loaded_signals, parse_errors = self._load_xlsx(filepath)

        except Exception as e:
            import traceback
            messagebox.showerror("Критическая ошибка загрузки", f"Произошла неожиданная ошибка при загрузке файла:\n{e}\n\n{traceback.format_exc()}", parent=self.master)
            return

        # --- Обработка результатов ---
        if parse_errors:
            error_details = "\n\n".join(parse_errors)
            # Укоротим сообщение, если оно слишком длинное
            if len(error_details) > 1000: error_details = error_details[:1000] + "\n..."
            messagebox.showerror("Ошибки загрузки", f"Обнаружены ошибки при разборе файла '{os.path.basename(filepath)}':\n\n{error_details}", parent=self.master)
            if not loaded_signals or not messagebox.askyesno("Продолжить?", "Несмотря на ошибки, загрузить успешно разобранные сигналы?", parent=self.master):
                return # Не загружаем ничего

        if not loaded_signals and not parse_errors:
             messagebox.showinfo("Файл пуст", f"Файл '{os.path.basename(filepath)}' не содержит данных сигналов или они не были распознаны.", parent=self.master)
             return

        # --- Обновление интерфейса ---
        self.signals = loaded_signals
        self.lb.delete(0, 'end')
        self._clear_points_table()
        self.update_from_table_btn.config(state='disabled')
        # self.updated_from_table_index = None # Больше не используется

        for s in self.signals:
            self.lb.insert('end', s.Name)

        if self.signals:
            self.lb.selection_set(0) # Выбираем первый сигнал
            self.on_select() # Обновляем поля и таблицу для первого сигнала
        else:
            self._clear_entries()
            self.draw() # Очистим график, если ничего не загружено

        # Рисуем загруженные сигналы (on_select вызовет draw)
        messagebox.showinfo("Загрузка завершена", f"Загружено {len(self.signals)} сигнал(ов) из файла:\n{os.path.basename(filepath)}.", parent=self.master)

    def _load_text(self, filepath) -> tuple[list[Signal], list[str]]:
        """Загружает сигналы из текстового файла (PULSE/PWL)."""
        signals = []
        errors = []
        try:
            with open(filepath, 'r', encoding='utf-8') as f: lines = f.readlines()
        except IOError as e:
            errors.append(f"Не удалось открыть файл: {e}")
            return signals, errors

        for lineno, line in enumerate(lines, 1):
            line = line.strip()
            if not line or line.startswith('#'): continue # Пропускаем пустые и комментарии

            try:
                if ':' not in line: raise ValueError("Отсутствует разделитель ':'")
                name_part, data_part = line.split(':', 1)
                name = name_part.strip()
                if not name: name = f"Signal_{lineno}" # Имя по умолчанию, если пустое

                data_part_upper = data_part.strip().upper() # Приводим к верхнему регистру для PULSE/PWL
                data_part_orig = data_part.strip() # Сохраняем оригинал для PWL парсинга

                if data_part_upper.startswith("PULSE(") and data_part_upper.endswith(")"):
                    inner = data_part_orig[len("PULSE("):-1].strip()
                    tokens = inner.split()
                    if len(tokens) != 8: raise ValueError(f"PULSE: Ожидалось 8 значений (V0 V1 TD TR TF TH TP N), получено {len(tokens)}")
                    # Парсим значения PULSE (V0 V1 TD TR TF TH TP N)
                    v0 = float(tokens[0])
                    v1 = float(tokens[1])
                    # Время парсим через parse_time, чтобы поддержать единицы
                    td = parse_time(tokens[2])
                    tr = parse_time(tokens[3])
                    tf = parse_time(tokens[4])
                    th = parse_time(tokens[5]) # Это TH из PULSE
                    tp_period = parse_time(tokens[6]) # Это T_PERIOD из PULSE
                    n = int(tokens[7])

                    if td < 0 or tr < 0 or tf < 0 or th < 0: raise ValueError("PULSE: Времена (TD, TR, TF, TH) не могут быть отрицательными")
                    if n < 0: raise ValueError("PULSE: Число импульсов (N) не может быть отрицательным")

                    # Рассчитываем TL из TP (периода)
                    # TP_period = TR + TH + TF + TL
                    tl = tp_period - (tr + th + tf)
                    if tl < -1e-12: # Допускаем небольшую погрешность вычислений
                         # Если TL отрицательный, возможно, данные некорректны
                         raise ValueError(f"PULSE: Вычисленное значение TL ({fmt(tl)}) отрицательно (TP={fmt(tp_period)}, TR={fmt(tr)}, TH={fmt(th)}, TF={fmt(tf)}). Проверьте параметры.")
                    else:
                        tl = max(0, tl) # Убедимся, что не отрицательное из-за погрешности

                    sig = Signal(V0=v0, V1=v1, TD=td, TR=tr, TF=tf, TH=th, TL=tl, N=n, Name=name)
                    # pwl_points не устанавливаем, сигнал будет генерироваться по параметрам
                    signals.append(sig)

                elif data_part_upper.startswith("PWL(") and data_part_upper.endswith(")"):
                    inner = data_part_orig[len("PWL("):-1].strip()
                    tokens = inner.split()
                    if not tokens: raise ValueError("PWL: Нет точек данных")
                    if len(tokens) % 2 != 0: raise ValueError("PWL: Нечетное число значений (ожидаются пары время-напряжение)")

                    pwl_times = []
                    pwl_voltages = []
                    last_t = -float('inf')
                    for i in range(0, len(tokens), 2):
                        t_str, v_str = tokens[i], tokens[i+1]
                        try:
                             # Время может быть без единиц (секунды) или с ними
                             t_val = parse_time(t_str)
                             v_val = float(v_str)     # Напряжение просто число
                        except ValueError as e_pair:
                             raise ValueError(f"PWL: Ошибка в паре '{t_str} {v_str}': {e_pair}")

                        if t_val < last_t - 1e-15: # Допуск для сравнения
                            raise ValueError(f"PWL: Время должно монотонно возрастать (нарушение: {fmt(t_val)} < {fmt(last_t)})")
                        # Допускаем одинаковое время для вертикальных фронтов, обновляя напряжение
                        if abs(t_val - last_t) < 1e-15 and pwl_times:
                            pwl_voltages[-1] = v_val # Обновляем напряжение последней точки
                        else:
                            pwl_times.append(t_val)
                            pwl_voltages.append(v_val)
                            last_t = t_val # Обновляем только если время реально изменилось

                    if len(pwl_times) < 1: raise ValueError("PWL: Необходимо минимум 1 точка данных")
                    # Если первая точка не t=0, добавим ее с напряжением первой точки
                    if pwl_times[0] > 1e-15:
                        pwl_times.insert(0, 0.0)
                        pwl_voltages.insert(0, pwl_voltages[0])

                    # Создаем "пустой" сигнал PULSE, т.к. не можем надежно восстановить параметры
                    v0_pwl = pwl_voltages[0]
                    v1_pwl = v0_pwl
                    for v in pwl_voltages:
                        if abs(v - v0_pwl) > 1e-9: v1_pwl = v; break

                    # Создаем базовый сигнал (параметры будут игнорироваться при отрисовке, т.к. есть pwl_points)
                    sig = Signal(V0=v0_pwl, V1=v1_pwl, TD=0, TR=1e-9, TF=1e-9, TH=0, TL=0, N=0, Name=name)
                    # Сразу устанавливаем PWL точки (set_pwl_points вызовет _update_time_scale_from_pwl)
                    sig.set_pwl_points(pwl_times, pwl_voltages)
                    signals.append(sig)

                else:
                    raise ValueError("Неизвестный формат: строка должна начинаться с 'Имя:PULSE(...)' или 'Имя:PWL(...)'")

            except Exception as ex:
                errors.append(f"Строка {lineno}: {line}\n  Ошибка: {ex}")

        return signals, errors

    def _parse_tabular_data(self, data_iterator, source_name) -> tuple[list[Signal], list[str]]:
        """Общий парсер для CSV и XLSX данных."""
        signals = []
        errors = []
        header = None
        signal_data = {} # {col_index: {'name': str, 'times': [], 'voltages': []}}
        time_col_index = -1

        for row_idx, row in enumerate(data_iterator):
            # Пропускаем пустые строки (особенно актуально для XLSX)
            if not any(cell is not None and str(cell).strip() != '' for cell in row):
                 continue

            if header is None: # Первая непустая строка - заголовок
                header = [str(h).strip() if h is not None else "" for h in row]
                # Ищем колонку времени (регистронезависимо)
                time_col_name = 'Time (s)'
                try:
                    # Ищем точное совпадение или регистронезависимое
                    found = False
                    for idx, h in enumerate(header):
                        if h == time_col_name or h.lower() == time_col_name.lower():
                            time_col_index = idx
                            found = True
                            break
                    if not found:
                         errors.append(f"Не найдена колонка времени '{time_col_name}' в заголовке: {header}")
                         return signals, errors # Критическая ошибка
                except ValueError:
                     errors.append(f"Не найдена колонка времени '{time_col_name}' в заголовке: {header}")
                     return signals, errors # Критическая ошибка

                # Инициализируем структуры для сигналов
                for col_idx, signal_name in enumerate(header):
                    if col_idx != time_col_index and signal_name: # Игнорируем колонку времени и пустые заголовки
                        signal_data[col_idx] = {'name': signal_name, 'times': [], 'voltages': []}
                if not signal_data:
                    errors.append("Не найдено ни одной колонки с именами сигналов в заголовке.")
                    return signals, errors
                continue # Переходим к следующей строке (данным)

            # --- Обработка строк данных ---
            if len(row) <= time_col_index:
                errors.append(f"Строка {row_idx + 1}: Недостаточно колонок для чтения времени (ожидалось {time_col_index + 1}, получено {len(row)}).")
                continue # Пропускаем строку

            time_val_raw = row[time_col_index]
            try:
                # Время в CSV/XLSX всегда в секундах (как при экспорте)
                time_val = float(time_val_raw)
                if time_val < 0:
                     raise ValueError("Время не может быть отрицательным")
            except (ValueError, TypeError) as e:
                errors.append(f"Строка {row_idx + 1}, колонка времени ({header[time_col_index]}): Неверное значение '{time_val_raw}' ({e}).")
                continue # Пропускаем строку, если время невалидно

            # Читаем напряжения для каждого сигнала
            for col_idx, sig_info in signal_data.items():
                if len(row) <= col_idx:
                    errors.append(f"Строка {row_idx + 1}: Недостаточно колонок для сигнала '{sig_info['name']}' (ожидалось {col_idx + 1}, получено {len(row)}).")
                    # Добавляем NaN или пропускаем? Пока пропускаем точку для этого сигнала
                    continue

                voltage_val_raw = row[col_idx]
                try:
                    voltage_val = float(voltage_val_raw)
                except (ValueError, TypeError) as e:
                    errors.append(f"Строка {row_idx + 1}, колонка '{sig_info['name']}': Неверное значение напряжения '{voltage_val_raw}' ({e}).")
                    # Используем NaN, чтобы обозначить пропуск? Или последнее известное значение?
                    # Пока просто пропустим добавление этой точки для этого сигнала
                    continue # Пропускаем эту точку для данного сигнала

                sig_info['times'].append(time_val)
                sig_info['voltages'].append(voltage_val)

        # --- Создание объектов Signal из собранных данных ---
        if not header:
            errors.append("Не найден заголовок в файле.")
            return signals, errors

        for col_idx, sig_info in signal_data.items():
            name = sig_info['name']
            times = sig_info['times']
            voltages = sig_info['voltages']

            if not times:
                errors.append(f"Сигнал '{name}': Не найдено валидных точек данных.")
                continue

            # Проверка монотонности времени (данные должны быть отсортированы по времени при экспорте)
            last_t = -float('inf')
            valid_times = []
            valid_voltages = []
            has_non_monotonic = False
            for t, v in zip(times, voltages):
                 if t < last_t - 1e-15:
                     has_non_monotonic = True
                     break # Нашли нарушение
                 # Допускаем одинаковое время - берем последнюю точку
                 if abs(t - last_t) < 1e-15 and valid_times:
                      valid_voltages[-1] = v
                 else:
                      valid_times.append(t)
                      valid_voltages.append(v)
                      last_t = t
            if has_non_monotonic:
                 errors.append(f"Сигнал '{name}': Данные времени не монотонно возрастают. Сигнал может быть некорректен.")
                 # Продолжаем с тем, что есть, но предупреждаем

            times = valid_times
            voltages = valid_voltages

            if not times: # Если после фильтрации ничего не осталось
                 errors.append(f"Сигнал '{name}': Не осталось валидных точек после проверки монотонности.")
                 continue

            # Если первая точка не t=0, добавим ее
            if times[0] > 1e-15:
                times.insert(0, 0.0)
                voltages.insert(0, voltages[0]) # Используем напряжение первой точки

            # Создаем Signal с PWL данными
            v0_pwl = voltages[0]
            v1_pwl = v0_pwl
            for v in voltages:
                if abs(v - v0_pwl) > 1e-9: v1_pwl = v; break

            sig = Signal(V0=v0_pwl, V1=v1_pwl, TD=0, TR=1e-9, TF=1e-9, TH=0, TL=0, N=0, Name=name)
            sig.set_pwl_points(times, voltages) # Устанавливаем PWL
            signals.append(sig)

        return signals, errors

    def _load_csv(self, filepath) -> tuple[list[Signal], list[str]]:
        """Загружает сигналы из CSV файла."""
        try:
            # Пробуем определить диалект CSV (разделитель)
            with open(filepath, 'r', newline='', encoding='utf-8') as csvfile:
                try:
                     dialect = csv.Sniffer().sniff(csvfile.read(1024*5)) # Читаем больше для надежности
                     csvfile.seek(0) # Возвращаемся к началу файла
                     reader = csv.reader(csvfile, dialect)
                     print(f"Detected CSV dialect: delimiter='{dialect.delimiter}', quotechar='{dialect.quotechar}'")
                except csv.Error:
                     # Если определить не удалось, используем стандартный разделитель ','
                     print("CSV dialect detection failed, using default delimiter=','")
                     csvfile.seek(0)
                     reader = csv.reader(csvfile, delimiter=',')

                # Передаем итератор reader в общий парсер
                return self._parse_tabular_data(reader, os.path.basename(filepath))

        except IOError as e:
            return [], [f"Не удалось прочитать CSV файл: {e}"]
        except Exception as e:
            import traceback
            return [], [f"Неожиданная ошибка при чтении CSV: {e}\n{traceback.format_exc()}"]


    def _load_xlsx(self, filepath) -> tuple[list[Signal], list[str]]:
        """Загружает сигналы из XLSX файла."""
        if not XLSX_SUPPORT: # Двойная проверка
            return [], ["Библиотека openpyxl не найдена."]
        try:
            wb = load_workbook(filename=filepath, read_only=True, data_only=True) # data_only=True для чтения значений формул
            # ws = wb.active # Берем активный лист
            # Или первый лист, что надежнее, если активный пустой
            ws = wb[wb.sheetnames[0]]
            # iter_rows(values_only=True) возвращает кортежи значений строк
            data_iterator = ws.iter_rows(values_only=True)
            return self._parse_tabular_data(data_iterator, os.path.basename(filepath))
        except FileNotFoundError:
            return [], [f"XLSX файл не найден: {filepath}"]
        except Exception as e:
            import traceback
            return [], [f"Ошибка при чтении XLSX файла: {e}\n{traceback.format_exc()}"]



# ───── Запуск приложения ─────
def main():
    try:
        root = tk.Tk()
        # Попробуем установить тему Windows для более нативного вида (если доступно)
        try:
            from tkinter import font
            default_font = font.nametofont("TkDefaultFont")
            default_font.configure(family="Segoe UI", size=10)
            root.option_add("*Font", default_font)
            ttk.Style().theme_use('vista') # или 'xpnative', 'winnative'
        except Exception:
             print("Windows theme not available, using default 'clam'.") # Оставляем clam по умолчанию

        app = PulseApp(root)
        root.mainloop()
    except Exception as e:
        import traceback
        # Используем tk окно для ошибки, если Tkinter еще работает
        try:
             root_err = tk.Tk()
             root_err.withdraw() # Скрываем основное окно
             messagebox.showerror("Критическая ошибка приложения", f"Произошла неперехваченная ошибка:\n{e}\n\n{traceback.format_exc()}", parent=None)
             root_err.destroy()
        except Exception as tk_err:
            # Если Tkinter совсем не работает, выводим в консоль
            print("Критическая ошибка приложения (Tkinter недоступен для отображения):")
            print(e)
            print(traceback.format_exc())
            print("\nОшибка Tkinter при попытке показать сообщение:")
            print(tk_err)


if __name__ == '__main__':
    main()
